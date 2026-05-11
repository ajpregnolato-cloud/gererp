"""Camada SQLite para os cadastros do Gerador CETESB.

Este módulo concentra a criação do banco local, importação inicial das planilhas
atuais e operações CRUD dos cadastros que antes ficavam em arquivos Excel ou no
próprio código.
"""

from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

DB_FILE = "gererp.db"


def caminho_banco(db_path: str | Path | None = None) -> Path:
    """Retorna o caminho absoluto do banco SQLite local."""
    if db_path is not None:
        return Path(db_path)
    return Path(__file__).resolve().with_name(DB_FILE)


def agora_iso() -> str:
    """Data/hora UTC em formato ISO para auditoria simples."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalizar_cnpj(valor: Any) -> str:
    """Mantém apenas números de um CNPJ/CPF informado."""
    return "".join(filter(str.isdigit, str(valor or "")))


def normalizar_placa(valor: Any) -> str:
    """Normaliza placas removendo caracteres não alfanuméricos."""
    return re.sub(r"[^0-9A-Z]", "", str(valor or "").upper())


def texto(valor: Any) -> str:
    """Converte valores de planilha para texto limpo, preservando códigos."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def conectar(db_path: str | Path | None = None) -> sqlite3.Connection:
    """Abre conexão com chaves estrangeiras habilitadas."""
    conn = sqlite3.connect(caminho_banco(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def inicializar_banco(db_path: str | Path | None = None) -> Path:
    """Cria as tabelas necessárias se ainda não existirem."""
    path = caminho_banco(db_path)
    with conectar(path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS cnpjs_cetesb (
                cnpj TEXT PRIMARY KEY,
                codigo_gerador_cnpjcpf TEXT NOT NULL,
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS residuos (
                codigo_interno TEXT PRIMARY KEY,
                codigo_ibama TEXT NOT NULL DEFAULT '',
                peso TEXT NOT NULL DEFAULT '',
                estado_fisico TEXT NOT NULL DEFAULT '',
                classe TEXT NOT NULL DEFAULT '',
                acondicionamento TEXT NOT NULL DEFAULT '',
                tratamento TEXT NOT NULL DEFAULT '',
                codigo_onu TEXT NOT NULL DEFAULT '',
                classe_risco TEXT NOT NULL DEFAULT '',
                nome_embarque TEXT NOT NULL DEFAULT '',
                grupo_embalagem TEXT NOT NULL DEFAULT '',
                descricao TEXT NOT NULL DEFAULT '',
                cod_res TEXT NOT NULL DEFAULT '',
                cod_trat TEXT NOT NULL DEFAULT '',
                usuario TEXT NOT NULL DEFAULT '',
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS placas_motoristas (
                placa TEXT PRIMARY KEY,
                motorista TEXT NOT NULL,
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS prefixos_residuos (
                prefixo TEXT NOT NULL,
                codigo_interno TEXT NOT NULL,
                ordem INTEGER NOT NULL DEFAULT 0,
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL,
                PRIMARY KEY (prefixo, codigo_interno),
                FOREIGN KEY (codigo_interno) REFERENCES residuos(codigo_interno)
                    ON UPDATE CASCADE ON DELETE RESTRICT
            );

            CREATE TABLE IF NOT EXISTS configuracoes (
                chave TEXT PRIMARY KEY,
                valor TEXT NOT NULL DEFAULT '',
                atualizado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS processamentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                arquivo_entrada TEXT NOT NULL DEFAULT '',
                arquivo_saida TEXT NOT NULL DEFAULT '',
                codigo_destinador TEXT NOT NULL DEFAULT '',
                codigo_transportador TEXT NOT NULL DEFAULT '',
                total_linhas_entrada INTEGER NOT NULL DEFAULT 0,
                total_linhas_saida INTEGER NOT NULL DEFAULT 0,
                total_avisos INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pendente',
                mensagem_erro TEXT NOT NULL DEFAULT '',
                criado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS processamento_avisos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                processamento_id INTEGER NOT NULL,
                tipo TEXT NOT NULL DEFAULT 'aviso',
                mensagem TEXT NOT NULL,
                linha_entrada INTEGER,
                cnpj TEXT NOT NULL DEFAULT '',
                documento TEXT NOT NULL DEFAULT '',
                placa TEXT NOT NULL DEFAULT '',
                criado_em TEXT NOT NULL,
                FOREIGN KEY (processamento_id) REFERENCES processamentos(id)
                    ON DELETE CASCADE
            );
            """
        )
    return path


def _cabecalhos(linha: Iterable[Any]) -> list[str]:
    return [texto(c).strip().lower() for c in linha]


def _indice(headers: list[str], termos: tuple[str, ...], padrao: int) -> int:
    return next((i for i, h in enumerate(headers) if any(t in h for t in termos)), padrao)


def salvar_cnpj(cnpj: Any, codigo_gerador_cnpjcpf: Any, db_path: str | Path | None = None) -> None:
    """Insere ou altera um CNPJ e seu código Gerador CNPJCPF."""
    cnpj_limpo = normalizar_cnpj(cnpj)
    codigo = texto(codigo_gerador_cnpjcpf)
    if not cnpj_limpo:
        raise ValueError("CNPJ obrigatório.")
    if not codigo:
        raise ValueError("Código Gerador CNPJCPF obrigatório.")

    agora = agora_iso()
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        conn.execute(
            """
            INSERT INTO cnpjs_cetesb
                (cnpj, codigo_gerador_cnpjcpf, ativo, criado_em, atualizado_em)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(cnpj) DO UPDATE SET
                codigo_gerador_cnpjcpf = excluded.codigo_gerador_cnpjcpf,
                ativo = 1,
                atualizado_em = excluded.atualizado_em
            """,
            (cnpj_limpo, codigo, agora, agora),
        )


def apagar_cnpj(cnpj: Any, db_path: str | Path | None = None, definitivo: bool = False) -> None:
    """Apaga um CNPJ; por padrão desativa para preservar histórico."""
    cnpj_limpo = normalizar_cnpj(cnpj)
    if not cnpj_limpo:
        raise ValueError("CNPJ obrigatório.")

    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        if definitivo:
            conn.execute("DELETE FROM cnpjs_cetesb WHERE cnpj = ?", (cnpj_limpo,))
        else:
            conn.execute(
                "UPDATE cnpjs_cetesb SET ativo = 0, atualizado_em = ? WHERE cnpj = ?",
                (agora_iso(), cnpj_limpo),
            )


def listar_cnpjs(db_path: str | Path | None = None, somente_ativos: bool = True) -> dict[str, str]:
    """Retorna CNPJ -> Código Gerador CNPJCPF."""
    inicializar_banco(db_path)
    sql = "SELECT cnpj, codigo_gerador_cnpjcpf FROM cnpjs_cetesb"
    if somente_ativos:
        sql += " WHERE ativo = 1"
    sql += " ORDER BY cnpj"
    with conectar(db_path) as conn:
        return {row["cnpj"]: row["codigo_gerador_cnpjcpf"] for row in conn.execute(sql)}


def salvar_residuo(dados: dict[str, Any], db_path: str | Path | None = None) -> None:
    """Insere ou altera um resíduo pelo código interno."""
    codigo_interno = texto(dados.get("codigo_interno") or dados.get("cod_interno"))
    if not codigo_interno:
        raise ValueError("Código interno do resíduo obrigatório.")

    campos = {
        "codigo_ibama": texto(dados.get("codigo_ibama") or dados.get("ibama")),
        "peso": texto(dados.get("peso")),
        "estado_fisico": texto(dados.get("estado_fisico")),
        "classe": texto(dados.get("classe")),
        "acondicionamento": texto(dados.get("acondicionamento")),
        "tratamento": texto(dados.get("tratamento")),
        "codigo_onu": texto(dados.get("codigo_onu") or dados.get("cod_onu")),
        "classe_risco": texto(dados.get("classe_risco")),
        "nome_embarque": texto(dados.get("nome_embarque")),
        "grupo_embalagem": texto(dados.get("grupo_embalagem")),
        "descricao": texto(dados.get("descricao")),
        "cod_res": texto(dados.get("cod_res")),
        "cod_trat": texto(dados.get("cod_trat")),
        "usuario": texto(dados.get("usuario")),
    }
    agora = agora_iso()
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        conn.execute(
            """
            INSERT INTO residuos (
                codigo_interno, codigo_ibama, peso, estado_fisico, classe,
                acondicionamento, tratamento, codigo_onu, classe_risco,
                nome_embarque, grupo_embalagem, descricao, cod_res, cod_trat,
                usuario, ativo, criado_em, atualizado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            ON CONFLICT(codigo_interno) DO UPDATE SET
                codigo_ibama = excluded.codigo_ibama,
                peso = excluded.peso,
                estado_fisico = excluded.estado_fisico,
                classe = excluded.classe,
                acondicionamento = excluded.acondicionamento,
                tratamento = excluded.tratamento,
                codigo_onu = excluded.codigo_onu,
                classe_risco = excluded.classe_risco,
                nome_embarque = excluded.nome_embarque,
                grupo_embalagem = excluded.grupo_embalagem,
                descricao = excluded.descricao,
                cod_res = excluded.cod_res,
                cod_trat = excluded.cod_trat,
                usuario = excluded.usuario,
                ativo = 1,
                atualizado_em = excluded.atualizado_em
            """,
            (codigo_interno, *campos.values(), agora, agora),
        )


def apagar_residuo(codigo_interno: Any, db_path: str | Path | None = None, definitivo: bool = False) -> None:
    """Apaga/desativa um resíduo e suas associações de prefixo."""
    codigo = texto(codigo_interno)
    if not codigo:
        raise ValueError("Código interno do resíduo obrigatório.")

    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        if definitivo:
            conn.execute("DELETE FROM prefixos_residuos WHERE codigo_interno = ?", (codigo,))
            conn.execute("DELETE FROM residuos WHERE codigo_interno = ?", (codigo,))
        else:
            agora = agora_iso()
            conn.execute(
                "UPDATE residuos SET ativo = 0, atualizado_em = ? WHERE codigo_interno = ?",
                (agora, codigo),
            )
            conn.execute(
                "UPDATE prefixos_residuos SET ativo = 0, atualizado_em = ? WHERE codigo_interno = ?",
                (agora, codigo),
            )


def listar_residuos(db_path: str | Path | None = None, somente_ativos: bool = True) -> dict[str, dict[str, str]]:
    """Retorna resíduos no formato esperado pelo gerador atual."""
    inicializar_banco(db_path)
    sql = "SELECT * FROM residuos"
    if somente_ativos:
        sql += " WHERE ativo = 1"
    sql += " ORDER BY codigo_interno"
    with conectar(db_path) as conn:
        residuos: dict[str, dict[str, str]] = {}
        for row in conn.execute(sql):
            residuos[row["codigo_interno"]] = {
                "ibama": row["codigo_ibama"],
                "classe": row["classe"],
                "descricao": row["descricao"],
                "estado_fisico": row["estado_fisico"],
                "acondicionamento": row["acondicionamento"],
                "tratamento": row["tratamento"],
                "cod_onu": row["codigo_onu"],
                "classe_risco": row["classe_risco"],
                "nome_embarque": row["nome_embarque"],
                "grupo_embalagem": row["grupo_embalagem"],
                "cod_interno": row["codigo_interno"],
            }
        return residuos


def listar_residuos_completos(db_path: str | Path | None = None, somente_ativos: bool = True) -> dict[str, dict[str, str]]:
    """Retorna todos os campos dos resíduos para telas de manutenção."""
    inicializar_banco(db_path)
    sql = "SELECT * FROM residuos"
    if somente_ativos:
        sql += " WHERE ativo = 1"
    sql += " ORDER BY codigo_interno"
    with conectar(db_path) as conn:
        return {row["codigo_interno"]: dict(row) for row in conn.execute(sql)}


def salvar_motorista(placa: Any, motorista: Any, db_path: str | Path | None = None) -> None:
    """Insere ou altera a relação placa -> motorista."""
    placa_norm = normalizar_placa(placa)
    nome = texto(motorista)
    if not placa_norm:
        raise ValueError("Placa obrigatória.")
    if not nome:
        raise ValueError("Motorista obrigatório.")

    agora = agora_iso()
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        conn.execute(
            """
            INSERT INTO placas_motoristas (placa, motorista, ativo, criado_em, atualizado_em)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(placa) DO UPDATE SET
                motorista = excluded.motorista,
                ativo = 1,
                atualizado_em = excluded.atualizado_em
            """,
            (placa_norm, nome, agora, agora),
        )


def apagar_motorista(placa: Any, db_path: str | Path | None = None, definitivo: bool = False) -> None:
    """Apaga/desativa uma placa cadastrada."""
    placa_norm = normalizar_placa(placa)
    if not placa_norm:
        raise ValueError("Placa obrigatória.")

    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        if definitivo:
            conn.execute("DELETE FROM placas_motoristas WHERE placa = ?", (placa_norm,))
        else:
            conn.execute(
                "UPDATE placas_motoristas SET ativo = 0, atualizado_em = ? WHERE placa = ?",
                (agora_iso(), placa_norm),
            )


def listar_motoristas(db_path: str | Path | None = None, somente_ativos: bool = True) -> dict[str, str]:
    """Retorna placa normalizada -> motorista."""
    inicializar_banco(db_path)
    sql = "SELECT placa, motorista FROM placas_motoristas"
    if somente_ativos:
        sql += " WHERE ativo = 1"
    with conectar(db_path) as conn:
        return {row["placa"]: row["motorista"] for row in conn.execute(sql)}


def salvar_prefixo_residuo(prefixo: Any, codigo_interno: Any, ordem: int = 0, db_path: str | Path | None = None) -> None:
    """Associa um prefixo de documento a um resíduo."""
    prefixo_txt = texto(prefixo)
    codigo = texto(codigo_interno)
    if not prefixo_txt:
        raise ValueError("Prefixo obrigatório.")
    if not codigo:
        raise ValueError("Código interno do resíduo obrigatório.")

    agora = agora_iso()
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        conn.execute(
            """
            INSERT INTO prefixos_residuos
                (prefixo, codigo_interno, ordem, ativo, criado_em, atualizado_em)
            VALUES (?, ?, ?, 1, ?, ?)
            ON CONFLICT(prefixo, codigo_interno) DO UPDATE SET
                ordem = excluded.ordem,
                ativo = 1,
                atualizado_em = excluded.atualizado_em
            """,
            (prefixo_txt, codigo, int(ordem), agora, agora),
        )


def apagar_prefixo_residuo(prefixo: Any, codigo_interno: Any, db_path: str | Path | None = None) -> None:
    """Remove a associação entre prefixo e resíduo."""
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        conn.execute(
            "DELETE FROM prefixos_residuos WHERE prefixo = ? AND codigo_interno = ?",
            (texto(prefixo), texto(codigo_interno)),
        )


def listar_prefixos_residuos(db_path: str | Path | None = None) -> dict[str, list[str]]:
    """Retorna prefixo -> lista ordenada de códigos internos de resíduos."""
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        grupos: dict[str, list[str]] = {}
        for row in conn.execute(
            """
            SELECT prefixo, codigo_interno
            FROM prefixos_residuos
            WHERE ativo = 1
            ORDER BY prefixo, ordem, codigo_interno
            """
        ):
            grupos.setdefault(row["prefixo"], []).append(row["codigo_interno"])
        return grupos


def importar_cnpjs_de_xlsx(arquivo: str | Path, db_path: str | Path | None = None) -> int:
    """Importa/atualiza CNPJs de uma planilha compatível com cnpj_cetesb.xlsx."""
    import openpyxl

    inicializar_banco(db_path)
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(max_row=1, values_only=True))
    headers = _cabecalhos(header)
    col_cnpj = _indice(headers, ("cnpj",), 0)
    col_cod = _indice(headers, ("gerador", "cnpjcpf"), 1)

    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(col_cnpj, col_cod):
            continue
        cnpj = normalizar_cnpj(row[col_cnpj])
        codigo = texto(row[col_cod])
        if cnpj and codigo:
            salvar_cnpj(cnpj, codigo, db_path)
            total += 1
    return total


def importar_residuos_de_xlsx(arquivo: str | Path, db_path: str | Path | None = None) -> int:
    """Importa/atualiza resíduos de uma planilha compatível com a listagem atual."""
    import datetime as dt
    import openpyxl

    inicializar_banco(db_path)
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active

    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        codigo_interno = texto(row[10] if len(row) > 10 else "")
        if not codigo_interno:
            continue

        classe = texto(row[3] if len(row) > 3 else "")
        ibama_raw = row[0] if len(row) > 0 else ""
        if isinstance(ibama_raw, dt.datetime):
            codigo_ibama = f"{ibama_raw.day:02d}{ibama_raw.month:02d}{str(ibama_raw.year)[2:]}"
        else:
            codigo_ibama = texto(ibama_raw).replace(" ", "")
        if classe == "I" and codigo_ibama and not codigo_ibama.endswith("(*)"):
            codigo_ibama += "(*)"

        salvar_residuo(
            {
                "codigo_interno": codigo_interno,
                "codigo_ibama": codigo_ibama,
                "peso": row[1] if len(row) > 1 else "",
                "estado_fisico": row[2] if len(row) > 2 else "",
                "classe": classe,
                "acondicionamento": row[4] if len(row) > 4 else "",
                "tratamento": texto(row[5] if len(row) > 5 else "").replace("\n", " "),
                "codigo_onu": "" if len(row) <= 6 or row[6] == "-" else row[6],
                "classe_risco": "" if len(row) <= 7 or row[7] == "-" else row[7],
                "nome_embarque": "" if len(row) <= 8 or row[8] == "-" else row[8],
                "grupo_embalagem": "" if len(row) <= 9 or row[9] == "-" else row[9],
                "descricao": row[11] if len(row) > 11 else "",
                "cod_res": row[12] if len(row) > 12 else "",
                "cod_trat": row[13] if len(row) > 13 else "",
                "usuario": row[14] if len(row) > 14 else "",
            },
            db_path,
        )
        total += 1
    return total


def importar_motoristas(motoristas: dict[str, str], db_path: str | Path | None = None) -> int:
    """Importa/atualiza o cadastro de placas e motoristas."""
    total = 0
    for placa, motorista in motoristas.items():
        salvar_motorista(placa, motorista, db_path)
        total += 1
    return total


def banco_tem_dados(db_path: str | Path | None = None) -> bool:
    """Indica se os cadastros principais já possuem dados."""
    inicializar_banco(db_path)
    with conectar(db_path) as conn:
        tabelas = ("cnpjs_cetesb", "residuos", "placas_motoristas")
        return any(conn.execute(f"SELECT EXISTS(SELECT 1 FROM {t} WHERE ativo = 1)").fetchone()[0] for t in tabelas)
