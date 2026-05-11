"""
Gerador CETESB por Prefixo
Dependências: pip install pandas openpyxl ttkbootstrap
"""

import threading, os, re, sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox

from database import (
    apagar_cnpj,
    apagar_motorista,
    apagar_residuo,
    importar_cnpjs_de_xlsx,
    importar_motoristas,
    importar_residuos_de_xlsx,
    inicializar_banco,
    listar_cnpjs,
    listar_motoristas,
    listar_prefixos_residuos,
    listar_residuos,
    listar_residuos_completos,
    salvar_cnpj,
    salvar_motorista,
    salvar_residuo,
    salvar_prefixo_residuo,
    apagar_prefixo_residuo,
)

# ── Arquivo externo de CNPJs ──────────────────────────────────────────────────
ARQUIVO_CNPJ = "cnpj_cetesb.xlsx"   # mesmo diretório do script

def carregar_cnpj_cetesb() -> dict:
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), ARQUIVO_CNPJ)
    try:
        inicializar_banco()
        if not listar_cnpjs(somente_ativos=False):
            if os.path.exists(caminho):
                total = importar_cnpjs_de_xlsx(caminho)
                print(f"✔ {total} CNPJs importados para o banco a partir de '{caminho}'")
            else:
                print(f"⚠ '{caminho}' não encontrado para importação inicial.")
        dados = listar_cnpjs()
        print(f"✔ {len(dados)} CNPJs carregados do banco local")
        return dados
    except Exception as e:
        print(f"⚠ Erro ao carregar CNPJs do banco: {e}")
        return {}

# ── Mapeamento de resíduos ────────────────────────────────────────────────────
ARQUIVO_RESIDUOS = "Listagem_Residuos_Geral_02_04_26.xlsx"

def carregar_residuos() -> dict:
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), ARQUIVO_RESIDUOS)
    try:
        inicializar_banco()
        if not listar_residuos(somente_ativos=False):
            if os.path.exists(caminho):
                total = importar_residuos_de_xlsx(caminho)
                print(f"✔ {total} resíduos importados para o banco a partir de '{caminho}'")
            else:
                print(f"⚠ '{caminho}' não encontrado para importação inicial.")
        residuos = listar_residuos()
        print(f"✔ {len(residuos)} resíduos carregados do banco local")
        return residuos
    except Exception as e:
        print(f"⚠ Erro ao carregar resíduos do banco: {e}")
        return {}

# ── Motoristas ────────────────────────────────────────────────────────────────
def _norm(p): return str(p).upper().replace("-","").replace(" ","").strip()

_MOT = {
    "FLI6673":"ROBSON","DAH7749":"JOSÉ APARECIDO","DAH7701":"ROGERIO",
    "DAH7738":"GRACIANO","DAH7751":"FLAVIO","DJB4194":"RUBENS",
    "EBM6241":"FABIANO DIAS","ELI5J48":"RODRIGO","ESU9D07":"DANILO",
    "FDC0611":"EDNILSON","FXP1C07":"SALVADOR","PZZ8C98":"FELIPE",
    "FFN2D31":"TERCEIRO","MOTOBOY":"TERCEIRO","NMO3E39":"TERCEIRO",
    "EGH0183":"TERCEIRO","EVU9D23":"TERCEIRO","DDU6647":"TERCEIRO",
    "CPN0060":"TERCEIRO","FDC0603":"OLIVEIRA","DAH7746":"ABEL",
    "DAH7737":"GRACIANO","DAH7743":"DIEGO","DAH7720":"FABIO",
    "DAH7744":"ROGERIO","DAH7742":"CELIO","GJN4484":"REGINALDO",
    "DAH7748":"JOZÉ BENTO","DAH7719":"RONALDO","DAH7712":"CELIO",
    "DAH7723":"FABIO","FVE6I75":"ROBSON","TER1518":"TERCEIRO",
    "FDC0605":"GILBERT0","FDC0602":"CELIO","FKC9G01":"LUIZ FERNANDO",
    "EZU2I27":"TERCEIRO","STF1G96":"TERCEIRO","TER0000":"TERCEIRO",
    "SFT1G93":"TERCEIRO","DAH7740":"CÉLIO","FAI4H92":"TERCEIRO",
    "DGH0183":"TERCEIRO","CPN0600":"TERCEIRO","CNP0060":"TERCEIRO",
    "OMG2G81":"TERCEIRO","OMG2681":"TERCEIRO","FRT7C02":"TERCEIRO",
    "NMO3E59":"TERCEIRO","DAH7797":"TERCEIRO","FDC0601":"DORIVAL",
    "EUL8I91":"TERCEIRO","EUL8191":"TERCEIRO","ESH7955":"ALTAIR",
    "EHS7955":"ALTAIR","EPO7718":"FELIPE CAMARGO","QRG5A18":"TERCEIRO",
    "GHYOI01":"TERCEIRO","ECO1H84":"TERCEIRO","DAH7722":"MARCOS",
    "BWI2566":"FERNANDO","QRL3D29":"TERCEIRO","SGC4J50":"TERCEIRO",
    "EFU1232":"JOSÉ APARECIDO","FDC0604":"FELIPE","DAH7747":"TERCEIRO",
    "LOJ5823":"BENEDITO","EUL9I91":"TERCEIRO","HEU4A92":"TERCEIRO",
    "HEU4A82":"TERCEIRO","SGC4J54":"TERCEIRO","SFT1G91":"TERCEIRO",
    "FJD8537":"TERCEIRO","FJD5837":"TERCEIRO","SFT1H84":"TERCEIRO",
    "STF1H84":"TERCEIRO","SCO1H84":"TERCEIRO","BWI2600":"ANDRÉ",
    "EZU9498":"FERNANDO","SUL9E04":"TERCEIRO","DKA8G77":"TERCEIRO",
    "DKA6G77":"TERCEIRO","GEK3F95":"TERCEIRO","DAH7727":"CESAR",
    "SSW9C62":"REGINALDO","GBE3A17":"TERCEIRO","ASB2E30":"FERNANDO",
    "FTB9975":"TERCEIRO","FWR2J93":"MARCOS LIMA","FTD9975":"TERCEIRO",
    "RYF5A06":"DIEGO","RYF5A08":"DIEGO","SUS5J49":"FELIPE",
    "FLG4071":"FABIANO","FWZ3I61":"LUIZ CESAR","TLP0G94":"PAULO",
    "EZU2127":"TERCEIRO","FXP1C07":"SALVADOR",
}
MOTORISTAS = {_norm(k): v for k, v in _MOT.items()}

def carregar_motoristas_banco() -> dict:
    try:
        inicializar_banco()
        if not listar_motoristas(somente_ativos=False):
            total = importar_motoristas(MOTORISTAS)
            print(f"✔ {total} motoristas importados para o banco local")
        dados = listar_motoristas()
        print(f"✔ {len(dados)} motoristas carregados do banco local")
        return dados
    except Exception as e:
        print(f"⚠ Erro ao carregar motoristas do banco: {e}")
        return MOTORISTAS

def get_motorista(placa): return MOTORISTAS.get(_norm(placa), "")
def limpar_cnpj(v): return "".join(filter(str.isdigit, str(v)))
def get_prefixo(doc):
    m = re.match(r"^(\d)", str(doc).strip())
    return m.group(1) if m else "?"

# ── Constantes ────────────────────────────────────────────────────────────────
CONTATO = "Antonio Jose Pregnolato"
UF      = "SP"

COLUNAS_SAIDA = [
    "Placa","UF","Motorista","Contato","CNPJ-Gerador","Código-CETESB",
    "Código-Destinador","Código-Transportador","Descrição","Resíduo",
    "Classe","Unidade-De-Medida","Tratamento","Estado-Físico",
    "Acondicionamento","Quantidade","Código-Interno","Nome-Interno",
    "Código-ONU","Classe-de-Risco","Nome-de-Embarque","Grupo-de-Embalagem",
]

COR_BG    = "#F0F2F5"
COR_CARD  = "#FFFFFF"
COR_PRIM  = "#1B4F8A"
COR_ACENT = "#2471A3"
COR_OK    = "#1A7A4A"
COR_ERR   = "#C0392B"
COR_MUTED = "#7F8C8D"
COR_BORDA = "#D5D8DC"
COR_TEXTO = "#1C2833"

# ── Geração do Excel ──────────────────────────────────────────────────────────
def gerar_excel(df_entrada, grupos_prefixo, cnpj_cetesb, residuos,
                cod_dest, cod_transp, caminho):
    linhas, avisos = [], []

    cols = list(df_entrada.columns)
    col_cnpj  = next((c for c in cols if "cnpj" in c.lower()), cols[0])
    col_doc   = next((c for c in cols if "doc" in c.lower() or "numero" in c.lower()), cols[1])
    col_placa = next((c for c in cols if "placa" in c.lower()), cols[2] if len(cols) > 2 else cols[0])

    for _, row in df_entrada.iterrows():
        cnpj_raw  = limpar_cnpj(str(row.get(col_cnpj, "")))
        doc       = str(row.get(col_doc, "")).strip()
        placa     = str(row.get(col_placa, "")).strip()
        prefixo   = get_prefixo(doc)
        res_ids   = grupos_prefixo.get(prefixo, [])

        if not res_ids:
            avisos.append(f"Doc '{doc}': prefixo '{prefixo}' sem resíduos → ignorado")
            continue

        cod_cetesb = cnpj_cetesb.get(cnpj_raw, "")
        if not cod_cetesb:
            avisos.append(f"CNPJ '{cnpj_raw}' sem Código-CETESB")

        for res_id in res_ids:
            res = residuos.get(res_id, {})
            if not res:
                avisos.append(f"Resíduo '{res_id}' não encontrado")
                continue

            classe_raw = res["classe"]
            classe_fmt = "CLASSE I" if classe_raw == "I" else "CLASSE IIA" if classe_raw == "II" else classe_raw

            linhas.append({
                "Placa":                placa,
                "UF":                   UF,
                "Motorista":            get_motorista(placa),
                "Contato":              CONTATO,
                "CNPJ-Gerador":         cnpj_raw,
                "Código-CETESB":        cod_cetesb,
                "Código-Destinador":    cod_dest,
                "Código-Transportador": cod_transp,
                "Descrição":            doc,
                "Resíduo":              res["ibama"],
                "Classe":               classe_fmt,
                "Unidade-De-Medida":    "Tonelada",
                "Tratamento":           res["tratamento"],
                "Estado-Físico":        res["estado_fisico"],
                "Acondicionamento":     res["acondicionamento"],
                "Quantidade":           0.001,
                "Código-Interno":       res["cod_interno"],
                "Nome-Interno":         res["descricao"],
                "Código-ONU":           "" if classe_raw == "II" else res["cod_onu"],
                "Classe-de-Risco":      "" if classe_raw == "II" else res["classe_risco"],
                "Nome-de-Embarque":     "" if classe_raw == "II" else res["nome_embarque"],
                "Grupo-de-Embalagem":   "" if classe_raw == "II" else res["grupo_embalagem"],
            })

    if not linhas:
        return [], avisos

    df = pd.DataFrame(linhas)[COLUNAS_SAIDA]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importação CETESB"

    h_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    h_fill  = PatternFill("solid", start_color="1B4F8A")
    c_font  = Font(name="Arial", size=10)
    c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin    = Side(style="thin", color="BFBFBF")
    brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, cn in enumerate(COLUNAS_SAIDA, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = c_align; cell.border = brd

    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for ci, cn in enumerate(COLUNAS_SAIDA, 1):
            valor = row_data[ci-1]
            cell  = ws.cell(row=ri, column=ci, value=valor if valor != "" else "")
            cell.font = c_font; cell.border = brd
            if cn == "Quantidade" and isinstance(valor, float):
                cell.number_format = "#,##0.0000"
                cell.alignment = c_align
            elif cn in {"CNPJ-Gerador","Código-CETESB","Classe","UF"}:
                cell.alignment = c_align
            else:
                cell.alignment = l_align

    larguras = {
        "Placa":12,"UF":6,"Motorista":18,"Contato":26,"CNPJ-Gerador":18,
        "Código-CETESB":16,"Código-Destinador":18,"Código-Transportador":20,
        "Descrição":20,"Resíduo":14,"Classe":12,"Unidade-De-Medida":14,
        "Tratamento":22,"Estado-Físico":14,"Acondicionamento":16,"Quantidade":12,
        "Código-Interno":16,"Nome-Interno":40,"Código-ONU":12,
        "Classe-de-Risco":14,"Nome-de-Embarque":40,"Grupo-de-Embalagem":18,
    }
    for ci, cn in enumerate(COLUNAS_SAIDA, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = larguras.get(cn, 16)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(caminho)
    return linhas, avisos

# ── Seletor de arquivo via PowerShell (evita bug tkinter 3.13/Windows) ────────
def selecionar_arquivo(modo="abrir", titulo="Selecionar", ext="xlsx"):
    import subprocess
    if sys.platform != "win32":
        from tkinter import filedialog
        if modo == "abrir":
            return filedialog.askopenfilename(title=titulo)
        return filedialog.asksaveasfilename(title=titulo, defaultextension=f".{ext}")

    if modo == "abrir":
        script = (
            "Add-Type -AssemblyName System.Windows.Forms;"
            "$f=New-Object System.Windows.Forms.OpenFileDialog;"
            f"$f.Title='{titulo}';"
            f"$f.Filter='Excel (*.{ext})|*.{ext}|Todos (*.*)|*.*';"
            "$f.ShowDialog()|Out-Null;Write-Output $f.FileName"
        )
    else:
        script = (
            "Add-Type -AssemblyName System.Windows.Forms;"
            "$f=New-Object System.Windows.Forms.SaveFileDialog;"
            f"$f.Title='{titulo}';"
            f"$f.Filter='Excel (*.{ext})|*.{ext}';"
            f"$f.DefaultExt='{ext}';"
            "$f.ShowDialog()|Out-Null;Write-Output $f.FileName"
        )
    try:
        r = subprocess.run(["powershell","-NoProfile","-Command",script],
                           capture_output=True, text=True, timeout=60)
        path = r.stdout.strip()
        return path if path else ""
    except Exception:
        return ""

# ── Janelas de manutenção dos cadastros ───────────────────────────────────────
class CadastroWindow:
    def __init__(self, app):
        self.app = app
        self.win = tk.Toplevel(app.root)
        self.win.title("Manutenção de Cadastros")
        self.win.geometry("980x680")
        self.win.configure(bg=COR_BG)
        self.win.transient(app.root)

        self.nb = ttk.Notebook(self.win)
        self.nb.pack(fill="both", expand=True, padx=10, pady=10)

        self._build_cnpjs()
        self._build_motoristas()
        self._build_residuos()

        self.win.protocol("WM_DELETE_WINDOW", self._fechar)

    def _tab(self, nome):
        frame = tk.Frame(self.nb, bg=COR_CARD)
        self.nb.add(frame, text=nome)
        return frame

    def _tree(self, parent, columns, headings, widths):
        box = tk.Frame(parent, bg=COR_CARD)
        box.pack(fill="both", expand=True, padx=10, pady=10)
        tree = ttk.Treeview(box, columns=columns, show="headings", height=12)
        ysb = ttk.Scrollbar(box, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(box, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        for col, head, width in zip(columns, headings, widths):
            tree.heading(col, text=head)
            tree.column(col, width=width, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        box.rowconfigure(0, weight=1)
        box.columnconfigure(0, weight=1)
        return tree

    def _form_entry(self, parent, row, label, var, width=34):
        tk.Label(parent, text=label, bg=COR_CARD, fg=COR_TEXTO,
                 font=("Segoe UI", 9, "bold"), anchor="w").grid(
            row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        ent = tk.Entry(parent, textvariable=var, width=width, bg="#F8F9FA",
                       relief="flat", highlightbackground=COR_BORDA,
                       highlightthickness=1, font=("Segoe UI", 9))
        ent.grid(row=row, column=1, sticky="ew", pady=4)
        return ent

    def _buttons(self, parent, novo, salvar, apagar):
        btns = tk.Frame(parent, bg=COR_CARD)
        btns.grid(row=99, column=0, columnspan=4, sticky="w", pady=(8, 0))
        tk.Button(btns, text="Novo/Limpar", bg=COR_ACENT, fg="white", relief="flat",
                  font=("Segoe UI", 9), cursor="hand2", padx=10, pady=3,
                  command=novo).pack(side="left")
        tk.Button(btns, text="Salvar", bg=COR_OK, fg="white", relief="flat",
                  font=("Segoe UI", 9), cursor="hand2", padx=10, pady=3,
                  command=salvar).pack(side="left", padx=6)
        tk.Button(btns, text="Apagar", bg=COR_ERR, fg="white", relief="flat",
                  font=("Segoe UI", 9), cursor="hand2", padx=10, pady=3,
                  command=apagar).pack(side="left")

    # CNPJs --------------------------------------------------------------------
    def _build_cnpjs(self):
        tab = self._tab("CNPJs")
        self.tree_cnpj = self._tree(
            tab, ("cnpj", "codigo"), ("CNPJ", "Gerador CNPJCPF"), (210, 160)
        )
        self.tree_cnpj.bind("<<TreeviewSelect>>", self._selecionar_cnpj)

        form = tk.Frame(tab, bg=COR_CARD)
        form.pack(fill="x", padx=10, pady=(0, 10))
        form.columnconfigure(1, weight=1)
        self.var_cad_cnpj = tk.StringVar()
        self.var_cad_codigo = tk.StringVar()
        self._form_entry(form, 0, "CNPJ", self.var_cad_cnpj)
        self._form_entry(form, 1, "Gerador CNPJCPF", self.var_cad_codigo)
        self._buttons(form, self._limpar_cnpj, self._salvar_cnpj_cadastro, self._apagar_cnpj_cadastro)
        self._refresh_cnpjs()

    def _refresh_cnpjs(self):
        for item in self.tree_cnpj.get_children():
            self.tree_cnpj.delete(item)
        for cnpj, codigo in listar_cnpjs().items():
            self.tree_cnpj.insert("", "end", values=(cnpj, codigo))

    def _selecionar_cnpj(self, _event=None):
        sel = self.tree_cnpj.selection()
        if not sel:
            return
        cnpj, codigo = self.tree_cnpj.item(sel[0], "values")
        self.var_cad_cnpj.set(cnpj)
        self.var_cad_codigo.set(codigo)

    def _limpar_cnpj(self):
        self.var_cad_cnpj.set("")
        self.var_cad_codigo.set("")
        self.tree_cnpj.selection_remove(self.tree_cnpj.selection())

    def _salvar_cnpj_cadastro(self):
        try:
            salvar_cnpj(self.var_cad_cnpj.get(), self.var_cad_codigo.get())
            self._refresh_cnpjs()
            self.app._recarregar_cadastros()
            self.app._log("✔ CNPJ salvo pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o CNPJ:\n{e}", parent=self.win)

    def _apagar_cnpj_cadastro(self):
        cnpj = limpar_cnpj(self.var_cad_cnpj.get())
        if not cnpj:
            messagebox.showwarning("Atenção", "Informe ou selecione um CNPJ.", parent=self.win)
            return
        if not messagebox.askyesno("Confirmar", f"Apagar/desativar o CNPJ {cnpj}?", parent=self.win):
            return
        try:
            apagar_cnpj(cnpj)
            self._limpar_cnpj()
            self._refresh_cnpjs()
            self.app._recarregar_cadastros()
            self.app._log("✔ CNPJ apagado pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível apagar o CNPJ:\n{e}", parent=self.win)

    # Motoristas ---------------------------------------------------------------
    def _build_motoristas(self):
        tab = self._tab("Motoristas")
        self.tree_motorista = self._tree(
            tab, ("placa", "motorista"), ("Placa", "Motorista"), (140, 260)
        )
        self.tree_motorista.bind("<<TreeviewSelect>>", self._selecionar_motorista)

        form = tk.Frame(tab, bg=COR_CARD)
        form.pack(fill="x", padx=10, pady=(0, 10))
        form.columnconfigure(1, weight=1)
        self.var_cad_placa = tk.StringVar()
        self.var_cad_motorista = tk.StringVar()
        self._form_entry(form, 0, "Placa", self.var_cad_placa)
        self._form_entry(form, 1, "Motorista", self.var_cad_motorista)
        self._buttons(form, self._limpar_motorista, self._salvar_motorista_cadastro, self._apagar_motorista_cadastro)
        self._refresh_motoristas()

    def _refresh_motoristas(self):
        for item in self.tree_motorista.get_children():
            self.tree_motorista.delete(item)
        for placa, motorista in sorted(listar_motoristas().items()):
            self.tree_motorista.insert("", "end", values=(placa, motorista))

    def _selecionar_motorista(self, _event=None):
        sel = self.tree_motorista.selection()
        if not sel:
            return
        placa, motorista = self.tree_motorista.item(sel[0], "values")
        self.var_cad_placa.set(placa)
        self.var_cad_motorista.set(motorista)

    def _limpar_motorista(self):
        self.var_cad_placa.set("")
        self.var_cad_motorista.set("")
        self.tree_motorista.selection_remove(self.tree_motorista.selection())

    def _salvar_motorista_cadastro(self):
        try:
            salvar_motorista(self.var_cad_placa.get(), self.var_cad_motorista.get())
            self._refresh_motoristas()
            self.app._recarregar_cadastros()
            self.app._log("✔ Motorista salvo pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o motorista:\n{e}", parent=self.win)

    def _apagar_motorista_cadastro(self):
        placa = self.var_cad_placa.get().strip()
        if not placa:
            messagebox.showwarning("Atenção", "Informe ou selecione uma placa.", parent=self.win)
            return
        if not messagebox.askyesno("Confirmar", f"Apagar/desativar a placa {placa}?", parent=self.win):
            return
        try:
            apagar_motorista(placa)
            self._limpar_motorista()
            self._refresh_motoristas()
            self.app._recarregar_cadastros()
            self.app._log("✔ Motorista apagado pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível apagar o motorista:\n{e}", parent=self.win)

    # Resíduos -----------------------------------------------------------------
    def _build_residuos(self):
        tab = self._tab("Resíduos")
        self.tree_residuo = self._tree(
            tab,
            ("codigo", "ibama", "classe", "descricao"),
            ("Código Interno", "IBAMA", "Classe", "Descrição"),
            (130, 120, 90, 520),
        )
        self.tree_residuo.bind("<<TreeviewSelect>>", self._selecionar_residuo)

        form = tk.Frame(tab, bg=COR_CARD)
        form.pack(fill="x", padx=10, pady=(0, 10))
        for col in (1, 3):
            form.columnconfigure(col, weight=1)
        self.res_vars = {campo: tk.StringVar() for campo in (
            "codigo_interno", "codigo_ibama", "peso", "estado_fisico", "classe",
            "acondicionamento", "tratamento", "codigo_onu", "classe_risco",
            "nome_embarque", "grupo_embalagem", "descricao", "cod_res",
            "cod_trat", "usuario",
        )}
        campos = [
            ("codigo_interno", "Código Interno"), ("codigo_ibama", "IBAMA"),
            ("peso", "Peso/Unidade"), ("estado_fisico", "Estado Físico"),
            ("classe", "Classe"), ("acondicionamento", "Acondicionamento"),
            ("tratamento", "Tratamento"), ("codigo_onu", "Código ONU"),
            ("classe_risco", "Classe de Risco"), ("grupo_embalagem", "Grupo Embalagem"),
            ("cod_res", "Cod Res"), ("cod_trat", "Cod Trat"),
            ("usuario", "Usuário"), ("nome_embarque", "Nome Embarque"),
            ("descricao", "Descrição"),
        ]
        for idx, (campo, label) in enumerate(campos):
            row = idx // 2
            base_col = (idx % 2) * 2
            tk.Label(form, text=label, bg=COR_CARD, fg=COR_TEXTO,
                     font=("Segoe UI", 9, "bold"), anchor="w").grid(
                row=row, column=base_col, sticky="w", padx=(0, 8), pady=3)
            width = 54 if campo in {"descricao", "nome_embarque"} else 28
            tk.Entry(form, textvariable=self.res_vars[campo], width=width, bg="#F8F9FA",
                     relief="flat", highlightbackground=COR_BORDA, highlightthickness=1,
                     font=("Segoe UI", 9)).grid(row=row, column=base_col+1, sticky="ew", pady=3)
        self._buttons(form, self._limpar_residuo, self._salvar_residuo_cadastro, self._apagar_residuo_cadastro)
        self._refresh_residuos()

    def _refresh_residuos(self):
        self.residuos_completos = listar_residuos_completos()
        for item in self.tree_residuo.get_children():
            self.tree_residuo.delete(item)
        for codigo, res in self.residuos_completos.items():
            self.tree_residuo.insert("", "end", values=(
                codigo, res.get("codigo_ibama", ""), res.get("classe", ""), res.get("descricao", "")
            ))

    def _selecionar_residuo(self, _event=None):
        sel = self.tree_residuo.selection()
        if not sel:
            return
        codigo = self.tree_residuo.item(sel[0], "values")[0]
        res = self.residuos_completos.get(codigo, {})
        for campo, var in self.res_vars.items():
            var.set(res.get(campo, ""))

    def _limpar_residuo(self):
        for var in self.res_vars.values():
            var.set("")
        self.tree_residuo.selection_remove(self.tree_residuo.selection())

    def _salvar_residuo_cadastro(self):
        try:
            salvar_residuo({campo: var.get() for campo, var in self.res_vars.items()})
            self._refresh_residuos()
            self.app._recarregar_cadastros()
            self.app._render_grupos()
            self.app._log("✔ Resíduo salvo pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o resíduo:\n{e}", parent=self.win)

    def _apagar_residuo_cadastro(self):
        codigo = self.res_vars["codigo_interno"].get().strip()
        if not codigo:
            messagebox.showwarning("Atenção", "Informe ou selecione um resíduo.", parent=self.win)
            return
        if not messagebox.askyesno("Confirmar", f"Apagar/desativar o resíduo {codigo}?", parent=self.win):
            return
        try:
            apagar_residuo(codigo)
            self._limpar_residuo()
            self._refresh_residuos()
            self.app._recarregar_cadastros()
            self.app._render_grupos()
            self.app._log("✔ Resíduo apagado pela tela de cadastros", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível apagar o resíduo:\n{e}", parent=self.win)

    def _fechar(self):
        self.app._recarregar_cadastros()
        self.win.destroy()


# ── Interface ─────────────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerador CETESB por Prefixo")
        self.root.geometry("700x1000")
        self.root.configure(bg=COR_BG)
        self.root.resizable(True, True)

        self.residuos        = {}
        self.cnpj_cetesb     = {}
        self.df_entrada      = None
        self.prefixos        = []
        self.grupos_prefixo  = {}
        self.tag_widgets     = {}

        self.var_cod_dest    = tk.StringVar()
        self.var_cod_transp  = tk.StringVar()
        self.var_arquivo     = tk.StringVar()
        self.var_saida       = tk.StringVar()
        self.var_cnpj_novo   = tk.StringVar()
        self.var_cetesb_novo = tk.StringVar()

        self._build()

        # Carrega dados externos em background
        threading.Thread(target=self._carregar_dados_bg, daemon=True).start()

    # ── Carregamento em background ────────────────────────────────────────────
    def _carregar_dados_bg(self):
        import time; time.sleep(0.4)
        global MOTORISTAS
        self.cnpj_cetesb = carregar_cnpj_cetesb()
        self.residuos    = carregar_residuos()
        MOTORISTAS       = carregar_motoristas_banco()
        self.grupos_prefixo.update(listar_prefixos_residuos())
        self.root.after(0, self._atualizar_apos_carga)

    def _atualizar_apos_carga(self):
        self._render_cnpj_resumo()
        self._log(f"✔ {len(self.cnpj_cetesb)} CNPJs  |  {len(self.residuos)} resíduos carregados", "ok")

    # ── Layout ────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self.root, bg=COR_PRIM, height=58)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="Gerador CETESB", bg=COR_PRIM, fg="white",
                 font=("Segoe UI",15,"bold")).pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Importação por Prefixo de Documento",
                 bg=COR_PRIM, fg="#A9CCE3", font=("Segoe UI",9)).pack(side="left")

        # Scroll
        outer = tk.Frame(self.root, bg=COR_BG)
        outer.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(outer, bg=COR_BG, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.fm = tk.Frame(self.canvas, bg=COR_BG)
        self.fm.bind("<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0,0), window=self.fm, anchor="nw")
        self.canvas.bind_all("<MouseWheel>",
            lambda e: self.canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._secao("1  Arquivo de entrada")
        self._sec_arquivo()
        self._secao("2  Campos fixos")
        self._sec_campos()
        self._secao("3  Cadastros do banco")
        self._sec_cadastros()
        self._secao("4  CNPJs carregados")
        self._sec_cnpj()
        self._secao("5  Grupos prefixo → resíduos")
        self.frame_grupos = self._card()
        tk.Label(self.frame_grupos, text="Carregue o arquivo de entrada para detectar prefixos.",
                 bg=COR_CARD, fg=COR_MUTED, font=("Segoe UI",9,"italic")).pack(
                 padx=12, pady=10, anchor="w")
        self._secao("6  Arquivo de saída")
        self._sec_saida()
        self._sec_botao()
        self._secao("Log")
        self._sec_log()

    def _secao(self, txt):
        f = tk.Frame(self.fm, bg=COR_BG)
        f.pack(fill="x", padx=20, pady=(14,3))
        tk.Label(f, text=txt, bg=COR_BG, fg=COR_PRIM,
                 font=("Segoe UI",10,"bold")).pack(side="left")
        tk.Frame(f, bg=COR_BORDA, height=1).pack(
            side="left", fill="x", expand=True, padx=(8,0), pady=6)

    def _card(self, pady=3):
        c = tk.Frame(self.fm, bg=COR_CARD,
                     highlightbackground=COR_BORDA, highlightthickness=1)
        c.pack(fill="x", padx=20, pady=pady)
        return c

    def _file_row(self, parent, lbl, var, cmd):
        f = tk.Frame(parent, bg=COR_CARD)
        f.pack(fill="x", padx=12, pady=8)
        tk.Label(f, text=lbl, bg=COR_CARD, fg=COR_TEXTO,
                 font=("Segoe UI",9,"bold"), width=34, anchor="w").pack(side="left")
        tk.Entry(f, textvariable=var, bg="#F8F9FA", relief="flat",
                 highlightbackground=COR_BORDA, highlightthickness=1,
                 font=("Segoe UI",9)).pack(side="left", fill="x", expand=True, padx=6)
        tk.Button(f, text="Procurar…", bg=COR_ACENT, fg="white",
                  relief="flat", font=("Segoe UI",9), cursor="hand2",
                  padx=10, pady=3, command=cmd).pack(side="left")

    def _sec_arquivo(self):
        c = self._card()
        self._file_row(c, "Planilha entrada  (CNPJ / Nº Documento / Placa):",
                       self.var_arquivo, self._browse_entrada)

    def _sec_campos(self):
        c = self._card()
        g = tk.Frame(c, bg=COR_CARD)
        g.pack(fill="x", padx=12, pady=8)
        for i, (lbl, var) in enumerate([("Código-Destinador", self.var_cod_dest),
                                         ("Código-Transportador", self.var_cod_transp)]):
            tk.Label(g, text=lbl, bg=COR_CARD, fg=COR_TEXTO,
                     font=("Segoe UI",9,"bold"), width=22, anchor="w").grid(
                row=i, column=0, sticky="w", pady=3)
            tk.Entry(g, textvariable=var, width=20, bg="#F8F9FA", relief="flat",
                     highlightbackground=COR_BORDA, highlightthickness=1,
                     font=("Segoe UI",9)).grid(row=i, column=1, sticky="w", padx=8)
        for i, (lbl, val) in enumerate([("Contato", CONTATO), ("UF", UF),
                                         ("Placa / Motorista", "Do arquivo de entrada")], 2):
            tk.Label(g, text=lbl, bg=COR_CARD, fg=COR_TEXTO,
                     font=("Segoe UI",9,"bold"), width=22, anchor="w").grid(
                row=i, column=0, sticky="w", pady=3)
            e = tk.Entry(g, width=40, bg="#EAECEE", relief="flat",
                         highlightbackground=COR_BORDA, highlightthickness=1,
                         font=("Segoe UI",9), fg=COR_MUTED)
            e.insert(0, val); e.configure(state="readonly")
            e.grid(row=i, column=1, sticky="w", padx=8)

    def _sec_cadastros(self):
        c = self._card()
        f = tk.Frame(c, bg=COR_CARD)
        f.pack(fill="x", padx=12, pady=(10, 4))
        tk.Label(f, text="Edite CNPJs, resíduos e placas/motoristas salvos no banco local.",
                 bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI",9)).pack(side="left")
        self.btn_cadastros = tk.Button(f, text="Abrir manutenção…", bg=COR_ACENT, fg="white",
                                       relief="flat", font=("Segoe UI",9), cursor="hand2",
                                       padx=12, pady=4, command=self._abrir_cadastros)
        self.btn_cadastros.pack(side="right")
        self.btn_import_residuos = tk.Button(
            f, text="Importar resíduos…", bg=COR_PRIM, fg="white",
            relief="flat", font=("Segoe UI",9), cursor="hand2",
            padx=12, pady=4, command=self._importar_residuos_planilha,
        )
        self.btn_import_residuos.pack(side="right", padx=(0, 8))
        self.btn_import_cnpjs = tk.Button(
            f, text="Importar CNPJs…", bg=COR_PRIM, fg="white",
            relief="flat", font=("Segoe UI",9), cursor="hand2",
            padx=12, pady=4, command=self._importar_cnpjs_planilha,
        )
        self.btn_import_cnpjs.pack(side="right", padx=(0, 8))

        prog = tk.Frame(c, bg=COR_CARD)
        prog.pack(fill="x", padx=12, pady=(0, 10))
        self.lbl_import_status = tk.Label(
            prog, text="Importação: aguardando seleção de planilha",
            bg=COR_CARD, fg=COR_MUTED, font=("Segoe UI",9), anchor="w",
        )
        self.lbl_import_status.pack(side="left")
        self.import_prog = ttk.Progressbar(prog, mode="determinate", maximum=100, length=220)
        self.import_prog.pack(side="right", padx=(8, 0))

    def _abrir_cadastros(self):
        CadastroWindow(self)

    def _importar_cnpjs_planilha(self):
        self._importar_planilha_auxiliar(
            tipo="CNPJs / Gerador CNPJCPF",
            arquivo_padrao=ARQUIVO_CNPJ,
            importador=importar_cnpjs_de_xlsx,
            atualizar_grupos=False,
        )

    def _importar_residuos_planilha(self):
        self._importar_planilha_auxiliar(
            tipo="resíduos",
            arquivo_padrao=ARQUIVO_RESIDUOS,
            importador=importar_residuos_de_xlsx,
            atualizar_grupos=True,
        )

    def _importar_planilha_auxiliar(self, tipo, arquivo_padrao, importador, atualizar_grupos=False):
        caminho = selecionar_arquivo(
            modo="abrir",
            titulo=f"Selecionar planilha de {tipo}",
            ext="xlsx",
        )
        if not caminho:
            return
        if not messagebox.askyesno(
            "Confirmar importação",
            f"Importar/atualizar {tipo} para o banco local?\n\n"
            f"Planilha selecionada:\n{caminho}\n\n"
            f"Arquivo esperado como referência: {arquivo_padrao}",
        ):
            return

        self._preparar_importacao(tipo)
        threading.Thread(
            target=self._thread_importar_planilha,
            args=(tipo, caminho, importador, atualizar_grupos),
            daemon=True,
        ).start()

    def _preparar_importacao(self, tipo):
        self._set_import_buttons_state("disabled")
        self.import_prog["value"] = 0
        self.lbl_import_status.configure(
            text=f"Importando {tipo}: 0% — iniciando leitura da planilha",
            fg=COR_ACENT,
        )
        self._log(f"⏳ Importando {tipo}...", "inf")

    def _set_import_buttons_state(self, state):
        for nome in ("btn_import_cnpjs", "btn_import_residuos", "btn_cadastros"):
            if hasattr(self, nome):
                getattr(self, nome).configure(state=state)

    def _thread_importar_planilha(self, tipo, caminho, importador, atualizar_grupos):
        try:
            def progress_callback(atual, total_linhas, importados):
                self.root.after(
                    0, self._atualizar_progresso_importacao,
                    tipo, atual, total_linhas, importados,
                )

            total = importador(caminho, progress_callback=progress_callback)
            self.root.after(
                0, self._finalizar_importacao_sucesso,
                tipo, caminho, total, atualizar_grupos,
            )
        except Exception as e:
            self.root.after(0, self._finalizar_importacao_erro, tipo, caminho, e)

    def _atualizar_progresso_importacao(self, tipo, atual, total_linhas, importados):
        total_linhas = max(total_linhas, 1)
        percentual = max(0, min(100, int((atual / total_linhas) * 100)))
        self.import_prog["value"] = percentual
        self.lbl_import_status.configure(
            text=(
                f"Importando {tipo}: {percentual}% "
                f"({atual}/{total_linhas} linhas, {importados} registros)"
            ),
            fg=COR_ACENT,
        )

    def _finalizar_importacao_sucesso(self, tipo, caminho, total, atualizar_grupos):
        self._recarregar_cadastros()
        if atualizar_grupos:
            self._render_grupos()
        self.import_prog["value"] = 100
        self.lbl_import_status.configure(
            text=f"Importação de {tipo}: 100% — {total} registro(s) importado(s)/atualizado(s)",
            fg=COR_OK,
        )
        self._set_import_buttons_state("normal")
        self._log(f"✔ Importação de {tipo} concluída: {total} registro(s) de {caminho}", "ok")
        messagebox.showinfo(
            "Importação concluída",
            f"Tipo importado: {tipo}\n"
            f"Registros importados/atualizados: {total}\n\n"
            f"Planilha:\n{caminho}",
        )

    def _finalizar_importacao_erro(self, tipo, caminho, erro):
        self._set_import_buttons_state("normal")
        self.lbl_import_status.configure(
            text=f"Importação de {tipo} falhou — verifique o log",
            fg=COR_ERR,
        )
        self._log(f"✖ Erro ao importar {tipo} de {caminho}: {erro}", "er")
        messagebox.showerror("Erro", f"Não foi possível importar {tipo}:\n{erro}")

    def _recarregar_cadastros(self):
        global MOTORISTAS
        self.cnpj_cetesb = listar_cnpjs()
        self.residuos = listar_residuos()
        MOTORISTAS = listar_motoristas()
        self.grupos_prefixo.update(listar_prefixos_residuos())
        if hasattr(self, "lbl_cnpj_total"):
            self._render_cnpj_resumo()

    def _sec_cnpj(self):
        c = self._card()
        # Resumo compacto — sem listar todos
        self.frame_cnpj_resumo = tk.Frame(c, bg=COR_CARD)
        self.frame_cnpj_resumo.pack(fill="x", padx=12, pady=(8,4))
        self.lbl_cnpj_total = tk.Label(self.frame_cnpj_resumo,
            text="Carregando CNPJs...", bg=COR_CARD, fg=COR_MUTED,
            font=("Segoe UI",9,"italic"))
        self.lbl_cnpj_total.pack(side="left")

        # Adicionar CNPJ manualmente
        add = tk.Frame(c, bg=COR_CARD)
        add.pack(fill="x", padx=12, pady=(0,8))
        tk.Label(add, text="Adicionar CNPJ:", bg=COR_CARD, fg=COR_TEXTO,
                 font=("Segoe UI",9,"bold")).pack(side="left")
        tk.Entry(add, textvariable=self.var_cnpj_novo, width=18, bg="#F8F9FA",
                 relief="flat", highlightbackground=COR_BORDA,
                 highlightthickness=1, font=("Segoe UI",9)).pack(side="left", padx=6)
        tk.Label(add, text="Gerador CNPJCPF:", bg=COR_CARD, fg=COR_TEXTO,
                 font=("Segoe UI",9,"bold")).pack(side="left")
        tk.Entry(add, textvariable=self.var_cetesb_novo, width=10, bg="#F8F9FA",
                 relief="flat", highlightbackground=COR_BORDA,
                 highlightthickness=1, font=("Segoe UI",9)).pack(side="left", padx=6)
        tk.Button(add, text="Salvar", bg=COR_OK, fg="white", relief="flat",
                  font=("Segoe UI",9), cursor="hand2", padx=8, pady=3,
                  command=self._adicionar_cnpj).pack(side="left")
        tk.Button(add, text="Apagar", bg=COR_ERR, fg="white", relief="flat",
                  font=("Segoe UI",9), cursor="hand2", padx=8, pady=3,
                  command=self._apagar_cnpj).pack(side="left", padx=(6,0))

    def _render_cnpj_resumo(self):
        total = len(self.cnpj_cetesb)
        self.lbl_cnpj_total.configure(
            text=f"✔ {total} CNPJs carregados do banco local  — salve ou apague cadastros abaixo",
            fg=COR_OK)

    def _adicionar_cnpj(self):
        cnpj = limpar_cnpj(self.var_cnpj_novo.get())
        cod  = self.var_cetesb_novo.get().strip()
        if not cnpj or not cod:
            messagebox.showwarning("Atenção", "Preencha CNPJ e Código Gerador CNPJCPF.")
            return
        try:
            salvar_cnpj(cnpj, cod)
            self.cnpj_cetesb = listar_cnpjs()
            self.var_cnpj_novo.set("")
            self.var_cetesb_novo.set("")
            self._render_cnpj_resumo()
            self._log(f"✔ CNPJ {cnpj} → {cod} salvo no banco", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o CNPJ:\n{e}")
            self._log(f"✖ Erro ao salvar CNPJ {cnpj}: {e}", "er")

    def _apagar_cnpj(self):
        cnpj = limpar_cnpj(self.var_cnpj_novo.get())
        if not cnpj:
            messagebox.showwarning("Atenção", "Informe o CNPJ que deseja apagar.")
            return
        if not messagebox.askyesno("Confirmar", f"Apagar/desativar o CNPJ {cnpj}?"):
            return
        try:
            apagar_cnpj(cnpj)
            self.cnpj_cetesb = listar_cnpjs()
            self.var_cnpj_novo.set("")
            self.var_cetesb_novo.set("")
            self._render_cnpj_resumo()
            self._log(f"✔ CNPJ {cnpj} apagado/desativado", "ok")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível apagar o CNPJ:\n{e}")
            self._log(f"✖ Erro ao apagar CNPJ {cnpj}: {e}", "er")

    def _sec_saida(self):
        c = self._card()
        self._file_row(c, "Salvar planilha CETESB como:", self.var_saida, self._browse_saida)

    def _sec_botao(self):
        f = tk.Frame(self.fm, bg=COR_BG)
        f.pack(fill="x", padx=20, pady=12)
        self.btn = tk.Button(f, text="▶  Gerar planilha CETESB",
                             bg=COR_OK, fg="white", relief="flat",
                             font=("Segoe UI",12,"bold"), cursor="hand2",
                             padx=22, pady=10, command=self._gerar)
        self.btn.pack(side="left")
        self.lbl_status = tk.Label(f, text="", bg=COR_BG, font=("Segoe UI",10))
        self.lbl_status.pack(side="left", padx=14)
        self.prog = ttk.Progressbar(f, mode="indeterminate", length=160)

    def _sec_log(self):
        c = self._card(pady=(0,16))
        self.log_txt = tk.Text(c, height=9, bg="#1C2833", fg="#A9DFBF",
                               font=("Consolas",9), relief="flat", state="disabled")
        self.log_txt.pack(fill="both", expand=True, padx=2, pady=2)
        self.log_txt.tag_config("ok", foreground="#58D68D")
        self.log_txt.tag_config("av", foreground="#F4D03F")
        self.log_txt.tag_config("er", foreground="#EC7063")
        self.log_txt.tag_config("inf",foreground="#85C1E9")

    # ── Grupos de prefixos ────────────────────────────────────────────────────
    def _render_grupos(self):
        for w in self.frame_grupos.winfo_children():
            w.destroy()
        self.tag_widgets = {}

        if not self.prefixos:
            tk.Label(self.frame_grupos,
                     text="Carregue o arquivo de entrada para detectar prefixos.",
                     bg=COR_CARD, fg=COR_MUTED,
                     font=("Segoe UI",9,"italic")).pack(padx=12, pady=10, anchor="w")
            return

        res_keys = list(self.residuos.keys())

        for prefixo in sorted(self.prefixos):
            bloco = tk.Frame(self.frame_grupos, bg="#F4F6F8",
                             highlightbackground=COR_BORDA, highlightthickness=1)
            bloco.pack(fill="x", padx=8, pady=5)

            hdr = tk.Frame(bloco, bg=COR_PRIM)
            hdr.pack(fill="x")
            n = len(self.grupos_prefixo.get(prefixo, []))
            tk.Label(hdr, text=f"  Prefixo  {prefixo}",
                     bg=COR_PRIM, fg="white",
                     font=("Segoe UI",10,"bold")).pack(side="left", pady=5)
            tk.Label(hdr, text=f"{n} resíduo(s)",
                     bg=COR_PRIM, fg="#A9CCE3",
                     font=("Segoe UI",9)).pack(side="left", padx=10)

            frame_tags = tk.Frame(bloco, bg="#F4F6F8")
            frame_tags.pack(fill="x", padx=8, pady=4)
            self.tag_widgets[prefixo] = frame_tags
            self._render_tags(prefixo)

            bot = tk.Frame(bloco, bg="#F4F6F8")
            bot.pack(fill="x", padx=8, pady=(0,6))
            tk.Label(bot, text="Adicionar:", bg="#F4F6F8", fg=COR_TEXTO,
                     font=("Segoe UI",9)).pack(side="left")

            opcoes_residuos = [self._formatar_residuo_opcao(k) for k in res_keys]
            var_sel = tk.StringVar()
            combo = ttk.Combobox(bot, textvariable=var_sel, width=65,
                                 font=("Segoe UI",9), state="normal")
            combo["values"] = opcoes_residuos
            combo.pack(side="left", padx=6)
            combo.bind("<KeyRelease>",
                       lambda _e, c=combo, ops=opcoes_residuos: self._filtrar_residuos_combo(c, ops))
            combo.bind("<FocusIn>",
                       lambda _e, c=combo, ops=opcoes_residuos: c.configure(values=ops))
            tk.Button(bot, text="+ Adicionar", bg=COR_OK, fg="white",
                      relief="flat", font=("Segoe UI",9), cursor="hand2",
                      padx=8, pady=2,
                      command=lambda p=prefixo, v=var_sel: self._add_residuo(p, v)
                      ).pack(side="left")

    def _formatar_residuo_opcao(self, res_id):
        res = self.residuos[res_id]
        return f"{res_id}  |  {res['ibama']}  —  {res['descricao'][:45]}"

    def _filtrar_residuos_combo(self, combo, opcoes):
        termo = combo.get().strip().lower()
        if termo:
            filtradas = [op for op in opcoes if termo in op.lower()]
        else:
            filtradas = opcoes
        combo.configure(values=filtradas)
        if filtradas:
            combo.event_generate("<Down>")

    def _render_tags(self, prefixo):
        frame = self.tag_widgets.get(prefixo)
        if not frame: return
        for w in frame.winfo_children(): w.destroy()
        ids = self.grupos_prefixo.get(prefixo, [])
        if not ids:
            tk.Label(frame, text="Nenhum resíduo adicionado",
                     bg="#F4F6F8", fg=COR_MUTED,
                     font=("Segoe UI",9,"italic")).pack(side="left")
            return
        for res_id in ids:
            res = self.residuos.get(res_id, {})
            tag = tk.Frame(frame, bg="#EBF5FB",
                           highlightbackground="#AED6F1", highlightthickness=1)
            tag.pack(side="left", padx=3, pady=2)
            tk.Label(tag, text=f"  {res.get('ibama', res_id)}",
                     bg="#EBF5FB", fg=COR_PRIM,
                     font=("Segoe UI",9,"bold")).pack(side="left", pady=2)
            tk.Button(tag, text=" ✕", bg="#EBF5FB", fg=COR_ERR,
                      relief="flat", font=("Segoe UI",8), cursor="hand2",
                      command=lambda p=prefixo, r=res_id: self._rem_residuo(p, r)
                      ).pack(side="left")

    def _add_residuo(self, prefixo, var_sel):
        val = var_sel.get()
        if not val: return
        res_id = val.split("  |  ")[0].strip()
        if res_id not in self.residuos:
            messagebox.showwarning("Atenção", "Selecione um resíduo válido da lista filtrada.")
            return
        lst = self.grupos_prefixo.setdefault(prefixo, [])
        if res_id not in lst:
            lst.append(res_id)
            try:
                salvar_prefixo_residuo(prefixo, res_id, len(lst))
                self._log(f"✔ Prefixo {prefixo} → resíduo {res_id} salvo", "ok")
            except Exception as e:
                self._log(f"⚠ Não foi possível salvar prefixo {prefixo}: {e}", "av")
        self._render_grupos()

    def _rem_residuo(self, prefixo, res_id):
        lst = self.grupos_prefixo.get(prefixo, [])
        if res_id in lst: lst.remove(res_id)
        try:
            apagar_prefixo_residuo(prefixo, res_id)
            self._log(f"✔ Prefixo {prefixo} → resíduo {res_id} removido", "ok")
        except Exception as e:
            self._log(f"⚠ Não foi possível remover prefixo {prefixo}: {e}", "av")
        self._render_grupos()

    # ── Browse ────────────────────────────────────────────────────────────────
    def _browse_entrada(self):
        path = selecionar_arquivo(modo="abrir", titulo="Selecionar planilha de entrada")
        if not path: return
        self.var_arquivo.set(path)
        self._carregar_prefixos(path)

    def _browse_saida(self):
        path = selecionar_arquivo(modo="salvar", titulo="Salvar planilha CETESB")
        if path: self.var_saida.set(path)

    def _carregar_prefixos(self, path):
        try:
            df = pd.read_excel(path, dtype=str).fillna("")
            df.columns = df.columns.str.strip()
            self.df_entrada = df
            cols = list(df.columns)
            col_doc = next((c for c in cols if "doc" in c.lower() or "numero" in c.lower()), cols[1])
            prefixos = sorted(set(get_prefixo(str(v)) for v in df[col_doc] if str(v).strip()))
            self.prefixos = prefixos
            for p in prefixos:
                self.grupos_prefixo.setdefault(p, [])
            self._render_grupos()
            self._log(f"✔ {len(df)} linhas  |  Prefixos: {', '.join(prefixos)}", "ok")
        except Exception as e:
            self._log(f"✖ Erro ao carregar: {e}", "er")

    # ── Geração ───────────────────────────────────────────────────────────────
    def _gerar(self):
        if self.df_entrada is None:
            messagebox.showwarning("Atenção", "Carregue o arquivo de entrada primeiro.")
            return
        if not self.var_cod_dest.get().strip():
            messagebox.showwarning("Atenção", "Preencha o Código-Destinador.")
            return
        if not self.var_cod_transp.get().strip():
            messagebox.showwarning("Atenção", "Preencha o Código-Transportador.")
            return
        if not self.var_saida.get().strip():
            messagebox.showwarning("Atenção", "Escolha o arquivo de saída.")
            return
        self.btn.configure(state="disabled", text="Processando…")
        self.prog.pack(side="left"); self.prog.start(10)
        self.lbl_status.configure(text="Gerando…", fg=COR_ACENT)
        self._log("─"*48)
        threading.Thread(target=self._thread_gerar, daemon=True).start()

    def _thread_gerar(self):
        try:
            linhas, avisos = gerar_excel(
                self.df_entrada, self.grupos_prefixo, self.cnpj_cetesb,
                self.residuos, self.var_cod_dest.get().strip(),
                self.var_cod_transp.get().strip(), self.var_saida.get().strip()
            )
            for av in avisos: self._log(f"⚠ {av}", "av")
            if linhas:
                self._log(f"✔ {len(linhas)} linhas geradas!", "ok")
                self._log(f"✔ Salvo em: {self.var_saida.get()}", "ok")
                self.root.after(0, lambda: messagebox.showinfo("Concluído!",
                    f"Planilha gerada!\n\n• {len(linhas)} linhas\n• {len(avisos)} avisos"))
                self.root.after(0, lambda: self.lbl_status.configure(
                    text=f"✔ {len(linhas)} linhas", fg=COR_OK))
            else:
                self._log("✖ Nenhuma linha gerada.", "er")
                self.root.after(0, lambda: self.lbl_status.configure(
                    text="Sem dados", fg=COR_ERR))
        except Exception as e:
            self._log(f"✖ {e}", "er")
            self.root.after(0, lambda: self.lbl_status.configure(text=f"Erro", fg=COR_ERR))
        finally:
            def _fim():
                self.prog.stop(); self.prog.pack_forget()
                self.btn.configure(state="normal", text="▶  Gerar planilha CETESB")
            self.root.after(0, _fim)

    def _log(self, msg, tag="inf"):
        def _do():
            self.log_txt.configure(state="normal")
            self.log_txt.insert("end", msg+"\n", tag)
            self.log_txt.see("end")
            self.log_txt.configure(state="disabled")
        self.root.after(0, _do)

def main():
    root = tk.Tk()
    try: root.iconbitmap("icon.ico")
    except: pass
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
